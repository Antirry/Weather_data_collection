import openpyxl
from _1_Rosstat_xlsx._1_Import_xlsx_rosstat import import_xlsx_rosstat as xlsx_name_path
import os


def check_file():
    try:
        file_name, current_path_xlsx = xlsx_name_path()
        file_path_name = current_path_xlsx + '\\' + file_name
        wb = openpyxl.load_workbook(filename=file_path_name)
        return wb
    except OSError:
        print("Нет файла или не то разрешение (должно быть .xlsx)")
        print("Ошибка в файле '_2_Read_xlsx.py'")
        return False


def maintenance_text_cells(wb) -> list[any]:
    hyperlinks_cities = list()
    text_cities = list()
    ws = wb['Содержание']

    for row in ws['C3':'C50']:
        for cell in row:
            try:
                hyperlinks_cities.append(cell.hyperlink.location)
                text_cities.append(cell.value)
            except:
                pass

    hyperlinks_cities = hyperlinks_cities[-5:-1]
    text_cities = text_cities[-5:-1]

    return [hyperlinks_cities, text_cities]


def check_list(text_hyperlinks_cities: list[any]) -> (list[str] | bool):
    text_cities = text_hyperlinks_cities[1]
    hyperlinks_cities = text_hyperlinks_cities[0]
    text_checked = 0

    text_cities_check = ['Города с численностью постоянного населения 1 млн. человек и более',
                         'Города с численностью постоянного населения от 500 тыс. до 1 млн. человек',
                         'Города с численностью постоянного населения от 250 тыс. до 500 тыс. человек',
                         'Города с численностью постоянного населения от 100 тыс. до 250 тыс. человек']

    print("ПРОВЕРКА \n\n", text_cities, "\n")

    for i in range(len(text_cities_check)):
        try:
            if text_cities[i] in text_cities_check[i]:
                text_checked += 1
        except Exception as ex:
            print("Сменилось положение в документе ссылок")
            print("Ошибка в файле '_2_Read_xlsx.py' -> ", ex)
            return False

    if text_checked != 0:
        print("УСПЕШНО \n")
        return hyperlinks_cities


def read_sheet_of_hyperlink(wb, hyperlinks_cities: list[str]) -> list[str]:
    cities = list()
    hyperlinks_cities = [hyperlink_city.replace('!A1', "") for hyperlink_city in hyperlinks_cities]

    for hyperlink_city in hyperlinks_cities:
        ws = wb[hyperlink_city]
        for row in ws['B7':'B50']:
            for cell in row:
                cities.append(cell.value)

    # Удаляю пометку о городе и пробелы после переменной (strip удаляет тег -> \n)
    cities = [city[2:].strip() for city in cities if city is not None and 'г ' in city]

    return cities[:50]


def read_xlsx() -> (list[str]):
    output = read_sheet_of_hyperlink(check_file(), check_list(maintenance_text_cells(check_file())))
    return output
