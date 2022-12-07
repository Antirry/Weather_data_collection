import openpyxl


def check_file():
    with open("Output.txt", "r") as f:
        line = f.readlines()
        # Получение названия файла
        line = line[0].strip()
        try:
            wb = openpyxl.load_workbook(line)
            f.close()
            return wb
        except OSError:
            f.close()
            print("Нет файла или не то разрешение (должно быть .xlsx)")
            return False


def maintenance_text_cells(wb) -> list[any]:
    hyperlinks_cities = list()
    text_cities = list()
    ws = wb['Содержание']

    for row in ws['C3':'C50']:
        for cell in row:
            try:
                hyperlinks_cities.append(cell.hyperlink.location)
                text_cities.append(cell.value)
            except AttributeError:
                pass

    hyperlinks_cities = hyperlinks_cities[-5:-1]
    text_cities = text_cities[-5:-1]

    return [hyperlinks_cities, text_cities]


def check_list(text_hyperlinks_cities: list[any]) -> (list[str] | bool):
    text_cities = text_hyperlinks_cities[1]
    hyperlinks_cities = text_hyperlinks_cities[0]

    text_cities_check = ['Города с численностью постоянного населения 1 млн. человек  и более',
                         'Города с численностью постоянного населения от 500 тыс. до 1 млн. человек',
                         'Города с численностью постоянного населения от 250 тыс. до 500 тыс. человек',
                         'Города с численностью постоянного населения от 100 тыс. до 250 тыс. человек']

    print("ПРОВЕРКА \n", text_cities_check)

    if text_cities == text_cities_check:
        print("\n\nУСПЕШНО\n\n")
        return hyperlinks_cities
    else:
        print("Сменилось положение в документе ссылок")
        return False


def read_sheet_of_hyperlink(wb, hyperlinks_cities: list[str]) -> list[str]:
    cities = list()
    hyperlinks_cities = [hyperlink_city.replace('!A1', "") for hyperlink_city in hyperlinks_cities]

    for hyperlink_city in hyperlinks_cities:
        ws = wb[hyperlink_city]
        for row in ws['B7':'B50']:
            for cell in row:
                cities.append(cell.value)

    cities = [city for city in cities if city is not None and 'г ' in city]

    return cities


def write_output_txt(output):
    with open("Output.txt", "a") as f:
        print(output, file=f)


def main():
    write_output_txt(read_sheet_of_hyperlink(check_file(), check_list(maintenance_text_cells(check_file()))))


main()